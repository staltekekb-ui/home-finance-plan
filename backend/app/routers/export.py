from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import date
from typing import Optional
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from app.database import get_db
from app.models import Transaction

router = APIRouter(prefix="/api/export", tags=["export"])


@router.get("/excel")
def export_to_excel(
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    category: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Экспорт транзакций в Excel файл."""

    query = db.query(Transaction)

    if date_from:
        query = query.filter(Transaction.date >= date_from)
    if date_to:
        query = query.filter(Transaction.date <= date_to)
    if category:
        query = query.filter(Transaction.category == category)

    transactions = query.order_by(Transaction.date.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Транзакции"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["Дата", "Описание", "Категория", "Сумма (₽)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row, t in enumerate(transactions, 2):
        ws.cell(row=row, column=1, value=t.date.strftime("%d.%m.%Y")).border = thin_border
        ws.cell(row=row, column=2, value=t.description).border = thin_border
        ws.cell(row=row, column=3, value=t.category or "—").border = thin_border
        amount_cell = ws.cell(row=row, column=4, value=t.amount)
        amount_cell.border = thin_border
        amount_cell.number_format = '#,##0.00'

    total_row = len(transactions) + 2
    ws.cell(row=total_row, column=3, value="Итого:").font = Font(bold=True)
    total_cell = ws.cell(row=total_row, column=4, value=sum(t.amount for t in transactions))
    total_cell.font = Font(bold=True)
    total_cell.number_format = '#,##0.00'

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "transactions"
    if date_from:
        filename += f"_from_{date_from}"
    if date_to:
        filename += f"_to_{date_to}"
    filename += ".xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
